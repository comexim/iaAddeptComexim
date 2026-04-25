"""
Serviço de envio de email via SMTP (Gmail)
"""
import io
import logging
import smtplib
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from typing import Optional

from app.core.config import settings

logger = logging.getLogger(__name__)


class EmailService:
    """Serviço de envio de email via SMTP"""

    def __init__(self):
        self.host = settings.smtp_host
        self.port = settings.smtp_port
        self.user = settings.smtp_user
        self.password = settings.smtp_password
        self.from_name = settings.smtp_from_name

    async def send_email(
        self,
        to: str,
        subject: str,
        body: str,
        xlsx_bytes: Optional[bytes] = None,
        xlsx_nome: str = "relatorio.xlsx",
    ) -> bool:
        """
        Envia email com corpo HTML e anexo xlsx opcional.

        Args:
            to: Email destinatário
            subject: Assunto do email
            body: Corpo do email (texto)
            xlsx_bytes: Planilha xlsx em bytes (opcional)
            xlsx_nome: Nome do arquivo xlsx

        Returns:
            True se enviado com sucesso
        """
        try:
            msg = MIMEMultipart("mixed")
            msg["Subject"] = subject
            msg["From"] = f"{self.from_name} <{self.user}>"
            msg["To"] = to

            # Parte alternativa (texto + html)
            alternativa = MIMEMultipart("alternative")
            alternativa.attach(MIMEText(body, "plain", "utf-8"))
            alternativa.attach(MIMEText(self._to_html(body), "html", "utf-8"))
            msg.attach(alternativa)

            # Anexo xlsx
            if xlsx_bytes:
                parte = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                parte.set_payload(xlsx_bytes)
                encoders.encode_base64(parte)
                parte.add_header("Content-Disposition", "attachment", filename=xlsx_nome)
                msg.attach(parte)

            with smtplib.SMTP(self.host, self.port) as server:
                server.starttls()
                server.login(self.user, self.password)
                server.sendmail(self.user, to, msg.as_string())

            sufixo = " (com planilha anexa)" if xlsx_bytes else ""
            logger.info(f"[EMAIL] Email enviado para {to} - assunto: {subject}{sufixo}")
            return True

        except Exception as e:
            logger.error(f"[EMAIL] Erro ao enviar email para {to}: {e}")
            return False

    @staticmethod
    def gerar_xlsx(data: list) -> Optional[bytes]:
        """Gera planilha xlsx formatada a partir de lista de dicts."""
        if not data:
            return None
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório"

            headers = list(data[0].keys())

            # Cabeçalho verde com texto branco negrito
            header_fill = PatternFill("solid", fgColor="2E7D32")
            header_font = Font(bold=True, color="FFFFFF")
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 20

            # Dados
            for row_idx, record in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=record.get(header))

            # Largura automática por coluna
            for col_idx, header in enumerate(headers, 1):
                letter = get_column_letter(col_idx)
                max_len = len(str(header))
                for record in data:
                    val = record.get(header)
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[letter].width = min(max_len + 3, 50)

            ws.freeze_panes = "A2"

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        except Exception as e:
            logger.error(f"[EMAIL] Erro ao gerar xlsx: {e}")
            return None

    def _to_html(self, text: str) -> str:
        """Converte texto com formatação WhatsApp (*negrito*) para HTML."""
        import re
        text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        text = re.sub(r"\*(.+?)\*", r"<strong>\1</strong>", text)
        text = text.replace("\n", "<br>")
        return f"""
        <html><body style="font-family: Arial, sans-serif; font-size: 14px; color: #333; max-width: 700px; margin: auto; padding: 20px;">
            <div style="background:#f5f5f5; border-left: 4px solid #2e7d32; padding: 16px; border-radius: 4px;">
                {text}
            </div>
            <p style="color:#999; font-size:12px; margin-top:20px;">
                Este é um relatório automático gerado pela Comexim IA.<br>
                Para cancelar, envie uma mensagem no WhatsApp.
            </p>
        </body></html>
        """


# Instância global
email_service = EmailService()
